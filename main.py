import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import CellIsRule
from tkinter.filedialog import askopenfilename, asksaveasfilename
import tkinter as tk
import tkinter.messagebox as msg

def read_namelist_cmd():
    """
    读取学生名单
    """
    global students, headers
    namelist_path = askopenfilename(
        filetypes=[('Excel Sheet', ('*.xlsx',))]
    )
    namelist_path_var.set(namelist_path)
    if not namelist_path:
        return
    workbook = openpyxl.load_workbook(namelist_path)
    sheet = workbook.active

    header_row = None
    headers = {}
    header_row_index = None
    for i, row in enumerate(sheet.iter_rows(
        min_row=1, 
        max_row=sheet.max_row
        )):
        row_values = [cell.value for cell in row]
        if ("姓名" in row_values) and (("学号" in row_values) or ("座位号" in row_values)):
            header_row = row_values
            header_row_index = i
            for idx, header in enumerate(header_row):
                if header == "姓名":
                    headers['name'] = idx
                if header in ["学号", "座位号"]:
                    headers['id'] = idx
                if header == "省份":
                    headers['province'] = idx
                if header == "学校":
                    headers['school'] = idx
            break
    if header_row is None:
        namelist_path_var.set('')
        namelist_text.config(state='normal')
        namelist_text.delete(1.0, tk.END)
        namelist_text.config(state='disabled')
        msg.showerror('错误', '未找到标题行，请检查工作表格式！')
        return
    
    students = []
    for row in sheet.iter_rows(min_row=header_row_index + 2, max_row=sheet.max_row):
        student_data = {}
        student_data['id'] = row[headers['id']].value
        student_data['name'] = row[headers['name']].value
        if 'province' in headers:
            student_data['province'] = row[headers['province']].value
        if 'school' in headers:
            student_data['school'] = row[headers['school']].value
        if student_data.get('id'):
            students.append(student_data)
    workbook.close()

    output = '学号: 姓名\n' + '\n'.join([f'{student["id"]}: {student["name"] if student["name"] else ""}' for student in students])

    namelist_text.config(state='normal')
    namelist_text.delete(1.0, tk.END)
    namelist_text.insert(tk.END, output)
    namelist_text.config(state='disabled')

    return

def read_scorelist_cmd():
    """
    读取分数分布（含百分比）
    """
    global n, scores_entries, rates_entries, scores, rates
    try:
        n = n_entry.get()
        assert n.isdigit()
        assert int(n) > 0
        n = int(n)
    except:
        msg.showerror('错误', '请输入一个合理的正整数！')
    
    for widget in table_frame.winfo_children():
        widget.destroy()
        scores_entries = []
        rates_entries = []
    
    # 第一行：题号
    for col in range(n):
        tk.Label(table_frame, text=f'{col+1}题').grid(row=0, column=col+1)
    tk.Label(table_frame, text='总和').grid(row=0, column=n+1)
    # 第二/三行：分数/占比
    tk.Label(table_frame, text='满分').grid(row=1, column=0)
    tk.Label(table_frame, text='占比').grid(row=2, column=0)
    score_sum = tk.Label(table_frame, text='0.0', foreground='green')
    score_sum.grid(row=1, column=n+1)
    rate_sum = tk.Label(table_frame, text='0.0', foreground='red')
    rate_sum.grid(row=2, column=n+1)
    for col in range(n):
        entry1 = tk.Entry(table_frame, width=4, textvariable=tk.DoubleVar(value=scores[col]) if scores[col] else tk.StringVar())
        entry2 = tk.Entry(table_frame, width=4, textvariable=tk.DoubleVar(value=rates[col]) if rates[col] else tk.StringVar())
        entry1.grid(row=1, column=col+1)
        entry2.grid(row=2, column=col+1)
        scores_entries.append(entry1)
        rates_entries.append(entry2)

        entry1.bind("<KeyRelease>", lambda event: update_score(score_sum, scores_entries, scores, 'score'))
        entry2.bind("<KeyRelease>", lambda event: update_score(rate_sum, rates_entries, rates, 'rate'))

    def update_score(widget, entries, values, name):
        total = 0
        for entry in entries:
            try:
                value = float(entry.get())
                total += value
            except:
                continue
        if name == 'rate':
            if total == 100:
                widget.config(text=f'{total:.1f}', foreground='green')
            else:
                widget.config(text=f'{total:.1f}', foreground='red')
        else:
            widget.config(text=f'{total:.1f}', foreground='green')
        for (i, entry) in enumerate(entries):
            try:
                value = float(entry.get())
                values[i] = value
            except:
                values[i] = 0
    pass

def generate_sheet_cmd():
    """
    生成分数表
    """
    try:
        assert students != []
        assert all(isinstance(_, float | int) for _ in scores[:n])
        assert 0 not in rates[:n]
        assert all(isinstance(_, float | int) for _ in rates[:n])
        assert sum(rates[:n]) == 100
    except:
        if students == []:
            msg.showerror('错误', '学生名单为空！')
        elif any(not isinstance(_, float | int) for _ in scores[:n]):
            msg.showerror('错误', '分数列表未填满！')
        elif 0 in rates[:n]:
            msg.showerror('错误', '存在满分为0的题！')
        elif any(not isinstance(_, float | int) for _ in rates[:n]):
            msg.showerror('错误', '分数列表未填满！')
        elif sum(rates[:n]) == 100:
            msg.showerror('错误', '占比总和不为100！')
        else:
            msg.showerror('错误', '未知错误')
        return
    
    workbook = openpyxl.Workbook()
    score_sheet = workbook.active
    headers = list(students[0].keys())
    curcol = 1
    currow = 1

    # 第一行：项目名
    indecies = {}
    if 'province' in headers:
        score_sheet[f'{col_order(curcol)}{currow}'] = '省份'
        score_sheet[f'{col_order(curcol)}{currow}'].font = Font(bold=True)
        score_sheet[f'{col_order(curcol)}{currow}'].alignment = Alignment(horizontal='center', vertical='center')
        indecies[curcol] = 'province'
        score_sheet.column_dimensions[col_order(curcol)].width = 8.2
        curcol += 1
    if 'school' in headers:
        score_sheet[f'{col_order(curcol)}{currow}'] = '学校'
        score_sheet[f'{col_order(curcol)}{currow}'].font = Font(bold=True)
        score_sheet[f'{col_order(curcol)}{currow}'].alignment = Alignment(horizontal='center', vertical='center')
        indecies[curcol] = 'school'
        score_sheet.column_dimensions[col_order(curcol)].width = 38.8
        curcol += 1
    score_sheet[f'{col_order(curcol)}{currow}'] = '姓名'
    score_sheet[f'{col_order(curcol)}{currow}'].font = Font(bold=True)
    score_sheet[f'{col_order(curcol)}{currow}'].alignment = Alignment(horizontal='center', vertical='center')
    indecies[curcol] = 'name'
    score_sheet.column_dimensions[col_order(curcol)].width = 8.2
    curcol += 1
    score_sheet[f'{col_order(curcol)}{currow}'] = '学号'
    score_sheet[f'{col_order(curcol)}{currow}'].font = Font(bold=True)
    score_sheet[f'{col_order(curcol)}{currow}'].alignment = Alignment(horizontal='center', vertical='center')
    indecies[curcol] = 'id'
    score_sheet.column_dimensions[col_order(curcol)].width = 8.2
    curcol += 1
    if 'province' in headers:
        score_sheet[f'{col_order(curcol)}{currow}'] = '本省人数'
        score_sheet[f'{col_order(curcol)}{currow}'].font = Font(bold=True)
        score_sheet[f'{col_order(curcol)}{currow}'].alignment = Alignment(horizontal='center', vertical='center')
        indecies[curcol] = 'count_p'
        curcol += 1
        score_sheet[f'{col_order(curcol)}{currow}'] = '本省排名'
        score_sheet[f'{col_order(curcol)}{currow}'].font = Font(bold=True)
        score_sheet[f'{col_order(curcol)}{currow}'].alignment = Alignment(horizontal='center', vertical='center')
        indecies[curcol] = 'rank_p'
        curcol += 1
    score_sheet[f'{col_order(curcol)}{currow}'] = '全场排名'
    score_sheet[f'{col_order(curcol)}{currow}'].font = Font(bold=True)
    score_sheet[f'{col_order(curcol)}{currow}'].alignment = Alignment(horizontal='center', vertical='center')
    indecies[curcol] = 'rank'
    curcol += 1

    score_sheet[f'{col_order(curcol)}{currow}'] = '总分'
    score_sheet[f'{col_order(curcol)}{currow}'].font = Font(bold=True)
    score_sheet[f'{col_order(curcol)}{currow}'].alignment = Alignment(horizontal='center', vertical='center')
    indecies[curcol] = 'total'
    total_col = curcol
    for i in range(n):
        curcol += 1
        score_sheet[f'{col_order(curcol)}{currow}'] = f'第{i+1}题'
        score_sheet[f'{col_order(curcol)}{currow}'].font = Font(bold=True)
        score_sheet[f'{col_order(curcol)}{currow}'].alignment = Alignment(horizontal='center', vertical='center')
        indecies[curcol] = i+1
    
    highlight = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    maxcol = curcol
    maxrow = len(students) + 1

    for i in range(n):
        curcol = total_col + i + 1
        rule = CellIsRule(operator='greaterThan', formula=[f'{scores[i]}'], fill=highlight)
        score_sheet.conditional_formatting.add(f'{col_order(curcol)}$2:{col_order(curcol)}${maxrow}', rule) 

    # 第二行起：填写学生信息及公式
    for student in students:
        currow += 1
        for curcol in range(1, maxcol+1):
            if indecies[curcol] in ['province', 'school', 'name', 'id']:
                score_sheet[f'{col_order(curcol)}{currow}'] = student[indecies[curcol]]
                score_sheet[f'{col_order(curcol)}{currow}'].alignment = Alignment(horizontal='center', vertical='center')
            elif indecies[curcol] == 'count_p':
                score_sheet[f'{col_order(curcol)}{currow}'] = f'=COUNTIF($A$2:$A${maxrow}, "*"&A{currow}&"*")'
                score_sheet[f'{col_order(curcol)}{currow}'].font = Font(bold=True)
                score_sheet[f'{col_order(curcol)}{currow}'].alignment = Alignment(horizontal='center', vertical='center')
            elif indecies[curcol] == 'rank_p':
                score_sheet[f'{col_order(curcol)}{currow}'] = f'=SUMPRODUCT(($A$2:$A${maxrow}=A{currow})*(${col_order(total_col)}$2:${col_order(total_col)}${maxrow}>{col_order(total_col)}{currow}))+1'
                score_sheet[f'{col_order(curcol)}{currow}'].font = Font(bold=True)
                score_sheet[f'{col_order(curcol)}{currow}'].alignment = Alignment(horizontal='center', vertical='center')
            elif indecies[curcol] == 'rank':
                score_sheet[f'{col_order(curcol)}{currow}'] = f'=RANK({col_order(curcol+1)}{currow}, {col_order(curcol+1)}$2:{col_order(curcol+1)}${maxrow})'
                score_sheet[f'{col_order(curcol)}{currow}'].font = Font(bold=True)
                score_sheet[f'{col_order(curcol)}{currow}'].alignment = Alignment(horizontal='center', vertical='center')
            elif indecies[curcol] == 'total':
                formula = '=' + '+'.join([f'{col_order(total_col+1+i)}{currow}*{rates[i]}/{scores[i]}' for i in range(n)])
                score_sheet[f'{col_order(curcol)}{currow}'] = formula
                score_sheet[f'{col_order(curcol)}{currow}'].font = Font(bold=True)
                score_sheet[f'{col_order(curcol)}{currow}'].alignment = Alignment(horizontal='center', vertical='center')
            elif isinstance(indecies[curcol], int):
                score_sheet[f'{col_order(curcol)}{currow}'].alignment = Alignment(horizontal='center', vertical='center')
                pass

    saveaspath = asksaveasfilename(defaultextension='*.xlsx', initialfile='分数表', filetypes=[('Excel Sheet', ('*.xlsx'))])
    try:
        workbook.save(saveaspath)
        msg.showinfo('成功', f'成功输出成绩表到{saveaspath}')
    except PermissionError:
        msg.showerror('错误', f'"{saveaspath}"文件已打开，请关闭后重试！')
    except:
        msg.showerror('错误', '未知错误')

def col_order(n: int):
    assert n > 0
    result = []
    while n > 0:
        n -= 1
        remainder = n % 26
        result.append(chr(65 + remainder))
        n //= 26
    return ''.join(result[::-1])


def clear_window():
    for widget in root.winfo_children():
        widget.destroy()

def window1():
    global namelist_path_entry, namelist_path_var, namelist_text
    clear_window()
    root.title('分数表生成器 | 读取学生名单')

    tk.Label(
        root,
        text='名单表：'
    ).place(x=70, y=30, anchor='e')
    namelist_path_entry = tk.Entry(
        root, 
        textvariable=namelist_path_var,
        width=52,
        state='disabled',
        background='white'
    )
    namelist_path_entry.place(x=70, y=30, anchor='w')
    tk.Button(
        root,
        text='打开',
        command=read_namelist_cmd
    ).place(x=450, y=30, anchor='w')
    namelist_text = tk.Text(
        root,
        height=50, width=40,
        bg='white',
    )
    namelist_text.pack(padx=10, pady=60)
    output = '学号: 姓名\n' + '\n'.join([f'{student["id"]}: {student["name"] if student["name"] else ""}' for student in students]) if students else '<名单预览>'
    namelist_text.delete(1.0, tk.END)
    namelist_text.insert(tk.END, output)
    namelist_text.config(state='disabled')

    tk.Button(
        root,
        text='下一步>>',
        width=30,
        command=window2
    ).place(x=250, y=470, anchor='c')
    
def window2():
    global n_entry, table_frame
    clear_window()
    root.title('分数表生成器 | 设置分数组成')

    tk.Button(
        root,
        text='<<返回',
        command=window1
    ).place(x=20, y=20)

    tk.Label(
        root,
        text='总题数：'
    ).place(x=30, y=60)

    n_entry = tk.Entry(
        root,
        textvariable=tk.IntVar(value=n),
        width=5
    )
    n_entry.place(x=80, y=60)
    tk.Button(
        root,
        text='刷新表格',
        command=read_scorelist_cmd
    ).place(x=120, y=55)

    table_frame = tk.Frame(root)
    table_frame.pack(pady=100)
    read_scorelist_cmd()

    tk.Button(
        root,
        text='生成成绩表>>',
        width=30,
        command=generate_sheet_cmd
    ).place(x=250, y=470, anchor='c')



if __name__ == "__main__":
    
    namelist_path=None
    students=[]

    root = tk.Tk()
    root.geometry('500x500')
    namelist_path_var = tk.StringVar()
    n = 10
    scores_entries = []
    rates_entries = []
    scores = ['']*20
    rates = ['']*20
    window1()
    root.mainloop()
    